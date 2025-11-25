# -*- coding:utf-8 -*-
##############################################################################
#
#   Módulo: odoo_import_invoice
#   Archivo: wizard/import_invoice.py
#   Descripción:
#       Asistente para importar facturas desde archivos Excel (.xls / .xlsx).
#       Permite crear facturas de clientes, proveedores, notas de crédito y débito.
#
#   FUNCIONALIDADES CLAVE:
#       ✅ Tipo de documento definido en el Excel (columna "invoice_type").
#       ✅ Diario contable tomado desde Excel.
#       ✅ Fecha contable soportada desde Excel.
#       ✅ Fecha de vencimiento soportada desde Excel ("invoice_date_due").
#       ✅ Cuenta contable configurable (desde producto o desde Excel).
#       ✅ Descarga de plantilla Excel con validaciones y traducciones dinámicas.
#       ✅ Validaciones de encabezados, fechas y valores numéricos.
#       ✅ Refactor modular, seguro y con manejo transaccional.
#       ✅ PARCHE: Corrección de propiedades del Workbook para evitar
#         mensajes de “reparación” en Microsoft Excel.
#
#   Autor: Allan E. Ramírez Madrid / INTEGRALL (2025)
#   Licencia: AGPL-3.0 or later (https://www.gnu.org/licenses/agpl)
#
##############################################################################

import base64
import xlrd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import datetime
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT

# ------------------------------------------------------------
# ENCABEZADOS ESPERADOS DEL EXCEL
# ------------------------------------------------------------
EXPECTED_HEADERS = [
    'excel_invoice', 'partner_vat_or_name', 'invoice_date', 'invoice_date_due',
    'product_code', 'description', 'quantity', 'unit_price', 'discount',
    'taxes', 'analytic_distribution', 'comment', 'ref_invoice',
    'tipo_gasto', 'currency', 'firma_fel', 'serie_fel', 'numero_fel',
    'journal_code', 'account_code', 'invoice_type', 'payment_ref', 'accounting_date'
]


class ImportInvoiceWizard(models.TransientModel):
    _name = 'import.invoice.wizard'
    _description = 'Import Invoice Wizard'

    # ------------------------------------------------------------
    # CAMPOS DEL ASISTENTE
    # ------------------------------------------------------------
    import_product_by = fields.Selection([
        ('name', 'Name'),
        ('code', 'Code'),
        ('barcode', 'Barcode')],
        string='Import Product By',
        default='name'
    )

    company_id = fields.Many2one(
        'res.company',
        required=True,
        string='Company',
        default=lambda self: self.env.company.id  # ✅ Compañía activa actual
    )

    invoice_stage_option = fields.Selection([
        ('draft', 'Import Draft Invoice'),
        ('validate', 'Validate Invoice Automatically with Import')],
        string='Invoice Stage Option',
        default='draft'
    )

    account_option = fields.Selection([
        ('product_incexp_account', 'Use Account from Configuration Product/Property'),
        ('from_excel_account', 'Use Account from Excel')],
        string='Account Option',
        default='from_excel_account'
    )

    files = fields.Binary(string="Import Excel File")
    datas_fname = fields.Char('Select Excel File')

    # ------------------------------------------------------------
    # MÉTODO PRINCIPAL DE IMPORTACIÓN
    # ------------------------------------------------------------
    def import_file(self):
        """Importar facturas desde un archivo Excel validado."""
        try:
            workbook = xlrd.open_workbook(file_contents=base64.decodebytes(self.files))
        except Exception:
            raise ValidationError(_("Please select a valid .xls or .xlsx file."))

        sheet = workbook.sheet_by_index(0)
        self._validate_headers(sheet)

        with self.env.cr.savepoint():
            created_moves = self._process_rows(sheet)

        return self._open_created_invoices(created_moves)

    # ------------------------------------------------------------
    # PROCESAR TODAS LAS FILAS DEL EXCEL
    # ------------------------------------------------------------
    def _process_rows(self, sheet):
        """Procesa cada fila del Excel y crea facturas agrupadas por excel_invoice."""
        currency_obj = self.env['res.currency']
        tax_obj = self.env['account.tax']
        product_obj = self.env['product.product']
        partner_obj = self.env['res.partner']
        product_account = self.env['account.account']

        number_of_rows = sheet.nrows
        invoice_dict = {}
        invoice_lists = []

        row = 1
        while row < number_of_rows:
            data = self._read_row(sheet, row)
            self._validate_row(data, row)

            # Resolución de datos principales
            partner = self._resolve_partner(data['partner_vat_or_name'], partner_obj)
            product = self._resolve_product(data['product_code'], product_obj)
            currency = currency_obj.search([('name', '=', data['currency'])], limit=1)
            if not currency:
                raise ValidationError(_("Invalid currency at row %s") % (row + 1))
            journal = self._resolve_journal(data['journal_code'])
            taxes = self._resolve_taxes(data['taxes'], tax_obj)
            account = self._resolve_account(data, product_account, product)

            # Fechas: emisión, vencimiento y contable
            invoice_date = self._parse_date(data['invoice_date'], row, "invoice_date")
            due_date = self._parse_date(data['invoice_date_due'], row, "invoice_date_due") if data[
                'invoice_date_due'] else None
            accounting_date = self._parse_date(data['accounting_date'] or data['invoice_date'], row, "accounting_date")

            # Crear cabecera si no existe
            if data['excel_invoice'] not in invoice_dict:
                vals = self._prepare_invoice_vals(
                    data, partner, journal, currency,
                    invoice_date, accounting_date, due_date
                )
                invoice_dict[data['excel_invoice']] = vals
                invoice_lists.append(data['excel_invoice'])
            else:
                vals = invoice_dict[data['excel_invoice']]

            # Agregar línea de factura
            line_vals = self._prepare_line_vals(data, product, taxes, account)
            vals['invoice_line_ids'].append((0, 0, line_vals))
            vals['narration'] = (vals['narration'] + ', ' + data['comment']).strip(', ')
            row += 1

        # Crear facturas en Odoo
        created_moves = self.env['account.move']
        for invoice in invoice_lists:
            move = self.env['account.move'].sudo().create(invoice_dict[invoice])
            if self.invoice_stage_option == 'validate':
                move.action_post()
            created_moves += move
        return created_moves

    # ------------------------------------------------------------
    # VALIDACIÓN Y LECTURA DE DATOS
    # ------------------------------------------------------------
    def _validate_headers(self, sheet):
        """Valida que todas las columnas esperadas existan en el Excel."""
        headers = [str(sheet.cell(0, c).value).strip() for c in range(sheet.ncols)]
        missing = [h for h in EXPECTED_HEADERS if h not in headers]
        if missing:
            raise ValidationError(_("Missing columns in Excel: %s") % ", ".join(missing))

    def _read_row(self, sheet, row):
        """Lee una fila del Excel y la transforma en un diccionario.

        ✅ Controla el tipo de interpretación por columna:
           - excel_invoice, partner_vat_or_name, quantity → número sin decimales
           - unit_price, discount → número con decimales
           - resto → texto (string)
        """
        cols = {name: idx for idx, name in enumerate(EXPECTED_HEADERS)}

        # Definición de columnas numéricas
        no_decimal_fields = ['excel_invoice', 'partner_vat_or_name', 'quantity', 'account_code']
        decimal_fields = ['unit_price', 'discount']

        def val(name):
            value = sheet.cell(row, cols[name]).value

            # Campo vacío
            if value in (None, ''):
                return ''

            # 🧮 Campos numéricos sin decimales
            if name in no_decimal_fields:
                if isinstance(value, float):
                    return str(int(value))  # Quita ".0"
                return str(value).strip()

            # 💲 Campos numéricos con decimales
            if name in decimal_fields:
                try:
                    return str(round(float(value), 6))  # Hasta 6 decimales para precios
                except Exception:
                    return '0'

            # 🔤 Resto: texto plano
            return str(value).strip()

        return {name: val(name) for name in EXPECTED_HEADERS}

    def _validate_row(self, data, row):
        """Valida los datos mínimos requeridos."""
        if not data['excel_invoice']:
            raise ValidationError(_("Missing invoice ID at row %s") % (row + 1))

        # ✅ CAMBIO: Ampliada la lista de tipos de movimiento válidos
        valid_invoice_types = [
            'out_invoice',  # Factura de Cliente
            'in_invoice',  # Factura de Proveedor
            'out_refund',  # Nota de Crédito Cliente
            'in_refund',  # Nota de Crédito Proveedor
            'out_receipt',  # Recibo de Cliente
            'in_receipt'  # Recibo de Proveedor
        ]
        if data['invoice_type'] not in valid_invoice_types:
            raise ValidationError(_("Invalid invoice_type at row %s. Expected one of: %s") % (
                row + 1, ', '.join(valid_invoice_types)
            ))

        if data['tipo_gasto'] not in ["mixto", "compra", "servicio", "importacion", "combustible"]:
            raise ValidationError(_("Invalid tipo_gasto at row %s") % (row + 1))

    def _parse_date(self, value, row, field):
        """Convierte texto a fecha en formato Odoo (YYYY-MM-DD)."""
        if not value:
            return False
        value = str(value).strip()
        for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).strftime(DEFAULT_SERVER_DATE_FORMAT)
            except Exception:
                continue
        raise ValidationError(_("Invalid date '%s' for %s at row %s (expected dd-mm-YYYY)") % (value, field, row))

    # ------------------------------------------------------------
    # RESOLVERS DE MODELOS
    # ------------------------------------------------------------
    def _resolve_partner(self, partner_value, partner_obj):
        """Busca el partner por NIT o nombre."""
        if not partner_value:
            raise ValidationError(_("Partner is empty"))
        partner = partner_obj.search([('vat', '=', partner_value)], limit=1)
        if not partner:
            partner = partner_obj.search([('name', '=', partner_value)], limit=1)
        if not partner:
            raise ValidationError(_("Partner not found: %s") % partner_value)
        return partner

    def _resolve_product(self, code, product_obj):
        """Obtiene el producto según la opción seleccionada en el wizard:
        - Si import_product_by = 'code' → busca en default_code.
        - Si import_product_by = 'name' → busca en name.
        - Si import_product_by = 'barcode' → busca en barcode.
        """
        if not code:
            raise ValidationError(_("Product code or name is missing."))

        search_domain = []
        search_label = ""

        if self.import_product_by == 'code':
            search_domain = [('default_code', '=', code)]
            search_label = _("internal code")
        elif self.import_product_by == 'name':
            search_domain = [('name', '=', code)]
            search_label = _("name")
        elif self.import_product_by == 'barcode':
            search_domain = [('barcode', '=', code)]
            search_label = _("barcode")
        else:
            raise ValidationError(_("Invalid product identification method."))

        product = product_obj.search(search_domain, limit=1)
        if not product:
            raise ValidationError(_("Product not found for %s: %s") % (search_label, code))
        return product

    def _resolve_journal(self, code):
        """Valida y obtiene el diario contable."""
        journal = self.env['account.journal'].search([
            ('code', '=', code),
            ('company_id', '=', self.company_id.id)
        ], limit=1)
        if not journal:
            raise ValidationError(_("Invalid journal code: %s") % code)
        return journal

    def _resolve_taxes(self, tax_str, tax_obj):
        """Convierte los nombres de impuestos en IDs."""
        tax_list = []
        if tax_str:
            names = [t.strip() for t in tax_str.split(',') if t.strip()]
            for tx in names:
                tax_id = tax_obj.search([
                    ('name', '=', tx),
                    ('company_id', '=', self.company_id.id)
                ], limit=1)
                if tax_id:
                    tax_list.append(tax_id.id)
        return tax_list

    def _resolve_account(self, data, product_account, product):
        """Determina la cuenta contable según la configuración del asistente."""
        if self.account_option == 'from_excel_account' and data.get('account_code'):
            acc = product_account.search([
                ('code', '=', data['account_code']),
                ('company_id', '=', self.company_id.id)
            ], limit=1)
            if not acc:
                raise ValidationError(_('Invalid account code "%s"') % data['account_code'])
            return acc
        else:
            move_type = self._get_move_type(data['invoice_type'])

            # ✅ CAMBIO: Ampliada la lógica de cuentas.
            # Tipos de cliente (out) usan cuentas de Ingreso.
            # Tipos de proveedor (in) usan cuentas de Gasto.
            if move_type in ['out_invoice', 'out_refund', 'out_receipt']:
                acc = product.property_account_income_id or product.categ_id.property_account_income_categ_id
            else:
                # in_invoice, in_refund, in_receipt
                acc = product.property_account_expense_id or product.categ_id.property_account_expense_categ_id

            if not acc:
                raise ValidationError(_('No valid account for product "%s"') % product.name)
            return acc

    # ------------------------------------------------------------
    # BUILDERS
    # ------------------------------------------------------------
    def _prepare_invoice_vals(self, data, partner, journal, currency, inv_date, acc_date, due_date):
        """Construye los valores del encabezado de la factura."""
        vals = {
            'partner_id': partner.id,
            'invoice_date': inv_date,
            'date': acc_date,
            'journal_id': journal.id,
            'company_id': self.company_id.id,
            'invoice_payment_term_id': partner.property_payment_term_id.id or False,
            'invoice_line_ids': [],
            'narration': '',
            'ref': data['ref_invoice'],
            'tipo_gasto': data['tipo_gasto'],
            'currency_id': currency.id,
            'firma_fel': data['firma_fel'],
            'serie_fel': data['serie_fel'],
            'numero_fel': data['numero_fel'],
            'payment_reference': data['payment_ref'],
            'move_type': self._get_move_type(data['invoice_type']),
        }
        if due_date:
            vals['invoice_date_due'] = due_date
        return vals

    def _prepare_line_vals(self, data, product, taxes, account):
        """Construye las líneas de factura."""
        return {
            'name': data['description'],
            'product_id': product.id,
            'quantity': float(data['quantity'] or 0),
            'discount': float(data['discount'] or 0),
            'price_unit': float(data['unit_price'] or 0),
            'tax_ids': [(6, 0, taxes)],
            'account_id': account.id,
        }

    # ------------------------------------------------------------
    # AUXILIARES
    # ------------------------------------------------------------
    def _get_move_type(self, invoice_type_excel):
        """Mapea el tipo de documento Excel a tipo de movimiento Odoo.

        ✅ CAMBIO: Los valores del Excel AHORA coinciden con los valores
        técnicos de Odoo (ej. 'out_invoice', 'in_receipt', etc.),
        los cuales ya fueron validados en _validate_row.
        """
        return invoice_type_excel

    def _open_created_invoices(self, created_moves):
        """Retorna acción para abrir las facturas creadas."""
        action = self.env["ir.actions.actions"]._for_xml_id("account.action_move_in_invoice_type")
        action['domain'] = [('id', 'in', created_moves.ids)]
        return action

    # ------------------------------------------------------------
    # DESCARGAR PLANTILLA DE EJEMPLO
    # ------------------------------------------------------------
    def download_template(self):
        """Genera y permite descargar una plantilla Excel de ejemplo.
        ✳️ Incluye soporte para traducción dinámica de filas de ayuda y ejemplo.
        ✅ Parche aplicado para eliminar advertencias de reparación en Excel.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = _("Invoice Import Template")

        # Encabezados con estilo visual
        header_fill = PatternFill(start_color="004D99", end_color="004D99", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(EXPECTED_HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Fila de ayuda (traducible)
        ws.append([
            _("Unique invoice ID"), _("Partner VAT or Name"), _("Invoice Date"), _("Due Date"),
            _("Product Code"), _("Description"), _("Qty"), _("Unit Price"), _("Disc %"),
            _("Tax names comma separated"), _("Analytic:Percent"), _("Comment"), _("Reference"),
            _("Expense type"), _("Currency"), _("FEL Signature"), _("FEL Serie"), _("FEL Number"),
            _("Journal Code"), _("Account Code"), _("Invoice Type"), _("Payment Ref"), _("Accounting Date")
        ])

        # ✅ CAMBIO: Definición de la lista de tipos de documento
        doc_types_list = [
            'out_invoice', 'in_invoice', 'out_refund', 'in_refund',
            'out_receipt', 'in_receipt'
        ]
        doc_types_str = ",".join(doc_types_list)

        # Fila de ejemplo (traducible)
        ws.append([
            'INV-001', _('41275748'), '15-10-2025', '30-10-2025', _('P0001'), _('Laptop Dell i7'), 1, 8500.00, 0,
            _('IVA 12%'), _('Ventas:100'), _('Primera importación'), _('FAC-001'), _('compra'),
            _('GTQ'), _('DD9822ED-671D-4D51-9181-6FAC9F406659'), _('DD9822ED'), _('1729973585'), _('COM'),
            _('51010101'),
            _('in_invoice'), _('Pago Contado'), '15-10-2025'  # ✅ CAMBIO: Ejemplo de tipo de documento
        ])

        # Validaciones de listas
        # ✅ CAMBIO: Validación de datos de Excel actualizada
        dv_type = DataValidation(type="list", formula1=f'"{doc_types_str}"', allow_blank=False)
        dv_gasto = DataValidation(type="list", formula1='"mixto,compra,servicio,importacion,combustible"',
                                  allow_blank=False)
        ws.add_data_validation(dv_type)
        ws.add_data_validation(dv_gasto)
        dv_type.add('U3:U500')
        dv_gasto.add('N3:N500')

        # Ajustar anchos de columnas automáticamente
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

        # ------------------------------------------------------------
        # PARCHE: Propiedades del workbook para evitar reparaciones
        # ------------------------------------------------------------
        wb.active = 0  # Define hoja activa
        ws.sheet_view.showGridLines = True
        if not wb.properties.title:
            wb.properties.title = "Invoice Import Template"
        wb.properties.creator = "Allan E. Ramirez Madrid https://www.integrall.solutions"
        wb.properties.lastModifiedBy = "Allan E. Ramirez Madrid https://www.integrall.solutions"
        wb.properties.created = datetime.now()
        wb.properties.modified = datetime.now()

        # Guardar en memoria y codificar
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)
        file_content = base64.b64encode(file_data.read())
        file_data.close()

        # Crear adjunto temporal
        attachment = self.env['ir.attachment'].create({
            'name': 'invoice_import_template.xlsx',
            'type': 'binary',
            'datas': file_content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        # Retornar acción de descarga directa
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }