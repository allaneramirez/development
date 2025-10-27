# -*- coding: utf-8 -*-
from odoo import models, fields, api, _  # Añadido _ para traducciones
from odoo.tools.misc import get_lang
import babel.dates
import io
import logging
import math # Añadido para comparación de floats

# --- IMPORTACIONES NECESARIAS ---
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter  # Importado para obtener letras de columna
except ImportError:
    openpyxl = None
# ------------------------------

_logger = logging.getLogger(__name__)


class AccountReport(models.Model):
    _inherit = 'account.report'

    show_sign_area = fields.Boolean(
        string="Show Sign Area",
        compute='_compute_show_sign_area'
    )

    def _compute_show_sign_area(self):
        # Asegúrate que 'account_reports.balance_sheet' sea el ID correcto
        financial_report = self.env.ref('account_reports.balance_sheet', raise_if_not_found=False)
        profit_and_loss = self.env.ref('account_reports.profit_and_loss', raise_if_not_found=False)

        for report in self:
            report_id = report.id
            is_financial_report = financial_report and report_id == financial_report.id
            is_profit_and_loss = profit_and_loss and report_id == profit_and_loss.id
            report.show_sign_area = is_financial_report or is_profit_and_loss

    def _format_date_es(self, date_obj):
        if not date_obj: return ''
        # Asegúrate de tener 'es_GT' instalado o usa 'es' como fallback
        lang_code = 'es_GT' if 'es_GT' in self.env['res.lang'].get_installed() else 'es'
        try:
            # Formato 'd de MMMM de y'
            formatted_date = babel.dates.format_date(date_obj, format='d MMMM y', locale=lang_code)
            parts = formatted_date.split(' ')
            if len(parts) == 3:
                # Reconstruir con 'de'
                return f"{parts[0]} de {parts[1]} de {parts[2]}"
            return formatted_date  # Fallback al formato por defecto si no tiene 3 partes
        except Exception as e:
            _logger.error(f"Error formateando fecha {date_obj} con locale {lang_code}: {e}")
            # Fallback a formatos sin locale si babel falla
            try:
                # Intenta formato español con strftime (depende del locale del servidor)
                return date_obj.strftime('%d de %B de %Y')
            except:
                # Fallback final a formato numérico
                return date_obj.strftime('%d/%m/%Y')

    def _get_report_rendering_context(self, options):
        """Prepara el contexto solo con nuestras variables personalizadas."""

        # 1. OBTENER SIEMPRE EL CONTEXTO BASE PRIMERO
        render_context = {}
        if hasattr(super(), '_get_report_rendering_context'):
            render_context = super()._get_report_rendering_context(options)
        else:
            # Fallback por si acaso
            render_context.update({
                'report': self,
                'options': options,
                'env': self.env,
                'report_company_name': self.env.company.display_name,
                'report_title': self.name,
            })

        # --- INICIO MODIFICACIÓN (ARREGLA LÓGICA HTML/PDF) ---
        # 2. VERIFICACIÓN CONDICIONAL
        # (Asumimos que self.show_sign_area fue calculado por get_html)
        if not self.show_sign_area:
            render_context['data'] = None  # Mantener consistencia
            return render_context  # Devolver contexto base SIN modificar
        # --- FIN MODIFICACIÓN ---

        # 3. SÓLO SI ES BALANCE O P&L, AÑADIR VARIABLES PERSONALIZADAS
        company = self.env.company
        report_name = self.name
        date_options = options.get('date', {})
        date_from_str = date_options.get('date_from')
        date_to_str = date_options.get('date_to')
        date_from = fields.Date.from_string(date_from_str) if date_from_str else None
        date_to = fields.Date.from_string(date_to_str) if date_to_str else None
        date_range_es = _("Periodo no especificado")  # Usar _() para traducción
        if date_from and date_to:
            date_range_es = _("Del %s al %s") % (self._format_date_es(date_from), self._format_date_es(date_to))
        elif date_to:
            date_range_es = _("Al %s") % (self._format_date_es(date_to))
        currency_name = company.currency_id.full_name or _('Moneda no especificada')
        currency_es = _("(Expresado en %s)") % currency_name  # Ahora sí usaremos esta variable

        # Actualizar el contexto base (o crearlo si no existe)
        render_context.update({
            'company_name': company.name or '',
            'company_vat': company.vat or '',
            'report_name': report_name or '',  # Nombre para nuestro encabezado
            'date_range_es': date_range_es,
            'currency_es': currency_es,  # La mantenemos en el contexto y ahora la usaremos
            'report': self,  # Asegurarse que 'report' esté
            'options': options,  # Asegurarse que 'options' esté
            'data': self.env.company if self.show_sign_area else None,
            # Mantener otros valores que podrías necesitar del contexto base si existía
            'env': self.env,
            'report_company_name': company.display_name,
            'report_title': self.name,
        })
        return render_context

    def get_html(self, options, *args, **kwargs):
        # --- INICIO MODIFICACIÓN (ARREGLA LÓGICA HTML/PDF) ---
        # Asegurarse de que 'show_sign_area' esté calculado
        # ANTES de llamar a _get_report_rendering_context.
        self.sudo()._compute_show_sign_area()
        # --- FIN MODIFICACIÓN ---

        local_context = self._get_report_rendering_context(options)
        original_additional_context = kwargs.get('additional_context', {})
        merged_additional_context = {**original_additional_context, **local_context}
        kwargs['additional_context'] = merged_additional_context
        html = super(AccountReport, self.with_context(discard_logo_check=True)).get_html(options, *args, **kwargs)
        if isinstance(html, bytes): html = html.decode('utf-8')
        return html

    def get_report_informations(self, options):
        info = super().get_report_informations(options)
        self.sudo()._compute_show_sign_area()
        info['show_sign_area'] = self.show_sign_area
        return info

    def export_to_xlsx(self, options, response=None):
        _logger.info(f"--- EJECUTANDO export_to_xlsx PARA REPORTE: {self.name} ---")

        # 1. OBTENER EL REPORTE ESTÁNDAR DE ODOO (ENTERPRISE)
        original_export_data = super(AccountReport, self.with_context(discard_logo_check=True)).export_to_xlsx(options,
                                                                                                               response)
        original_output_bytes = original_export_data['file_content']
        self.sudo()._compute_show_sign_area()
        _logger.info(f"--- Valor de show_sign_area: {self.show_sign_area} ---")

        if not self.show_sign_area:
            _logger.info("Reporte no configurado para firmas, devolviendo XLSX estándar.")
            return original_export_data
        if not openpyxl:
            _logger.warning("La librería 'openpyxl' no está instalada. Devolviendo XLSX estándar.")
            return original_export_data

        _logger.info(f"--- MODIFICANDO XLSX CON ENCABEZADO Y FIRMAS (reporte: {self.name}) ---")
        original_output_stream = io.BytesIO(original_output_bytes)
        try:
            # 2. CARGAR EL REPORTE ESTÁNDAR EN MEMORIA
            workbook = openpyxl.load_workbook(original_output_stream)
            try:
                sheet = workbook.active
            except Exception:
                sheet = workbook.worksheets[0]
            company = self.env.company
            bold_font = Font(name='Arial', size=12, bold=True)
            normal_font = Font(name='Arial', size=11)

            # --- INICIO MODIFICACIÓN (REQUISITO EXCEL CORREGIDO) ---
            # Fuente para las líneas de Nivel 1 (ahora tamaño 12)
            bold_data_font = Font(name='Arial', size=12, bold=True)
            # --- FIN MODIFICACIÓN ---

            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            sign_center_align = Alignment(horizontal='center', vertical='bottom')

            # --- 3. AGREGAR ENCABEZADO PERSONALIZADO ---
            # (Esta sección no cambia)
            render_context = self._get_report_rendering_context(options)
            _logger.info(f"--- Variables de contexto para encabezado ---")
            # (...) logs omitidos
            header_lines = []
            if render_context.get('company_name'): header_lines.append(str(render_context.get('company_name', '')).upper())
            if render_context.get('company_vat'): header_lines.append(f"NIT: {render_context.get('company_vat', '')}")
            if render_context.get('report_name'): header_lines.append(str(render_context.get('report_name', '')).upper())
            if render_context.get('date_range_es'): header_lines.append(str(render_context.get('date_range_es', '')))
            if render_context.get('currency_es'): header_lines.append(str(render_context.get('currency_es', '')))
            header_text = "\n".join(filter(lambda x: x is not None, header_lines))
            _logger.info(f"--- Texto final del encabezado para Excel: ---\n{header_text}")
            if header_text:
                sheet.insert_rows(1)
                num_lines_in_header = header_text.count('\n') + 1
                new_height = max(20, num_lines_in_header * 15)
                if 1 in sheet.row_dimensions: sheet.row_dimensions[1].height = new_height
                else: sheet.row_dimensions[1] = openpyxl.worksheet.dimensions.RowDimension(sheet, index=1, ht=new_height)
                _logger.info(f"Ajustando altura de fila 1 a: {new_height}")
                max_col = sheet.max_column if sheet.max_column >= 1 else 1
                max_col_letter = get_column_letter(max_col)
                merge_range = f'A1:{max_col_letter}1'
                try: sheet.merge_cells(merge_range)
                except Exception as merge_err:
                    _logger.error(f"Error al combinar celdas para encabezado '{merge_range}': {merge_err}. Intentando con A1.")
                    merge_range = 'A1'
                cell = sheet['A1']
                cell.value = header_text
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # =========================================================================
            # --- 4. ELIMINAR FILAS DE ENCABEZADO ORIGINALES (LÓGICA MEJORADA) ---
            # =========================================================================
            _logger.info("Buscando inicio del cuerpo del reporte (ej. 'ACTIVOS', 'INGRESOS')...")

            # --- INICIO MODIFICACIÓN (REQUISITO EXCEL CORREGIDO) ---
            # Palabras clave de TÍTULOS PRINCIPALES (Nivel 0) para encontrar el inicio
            body_start_keywords = (
                "activos", "assets",
                "pasivos", "liabilities",
                "capital", "equity",
                "ganancias netas", "net profit",
                "ingresos", "income",
                "gastos", "expenses"
            )
            # Tupla para la comprobación exacta de inserción de fila (y comparación insensible a mayúsculas)
            SPECIAL_HEADERS_TUPLE = (
                'ACTIVOS', 'PASIVOS', 'CAPITAL', 'PASIVO Y CAPITAL', 'PASIVO + CAPITAL',
                'GANANCIAS NETAS', 'INGRESOS', 'GASTOS',
                'ASSETS', 'LIABILITIES', 'EQUITY', 'LIABILITIES + EQUITY', 'NET PROFIT', 'INCOME', 'EXPENSES'
            )
            # --- FIN MODIFICACIÓN ---

            first_data_row_index = -1
            max_rows_to_check = 10
            first_data_row_actual_index = 2

            for row_idx in range(2, min(sheet.max_row + 1, 2 + max_rows_to_check)):
                cell_value = sheet.cell(row=row_idx, column=1).value
                if isinstance(cell_value, str):
                    cell_text = cell_value.strip().lower()
                    if any(keyword in cell_text for keyword in body_start_keywords):
                        _logger.info(f"Cuerpo del reporte encontrado en fila {row_idx} (contenido: '{cell_value}')")
                        first_data_row_index = row_idx
                        first_data_row_actual_index = row_idx
                        break

            if first_data_row_index > 2:
                rows_to_delete_count = first_data_row_index - 2
                _logger.info(f"Eliminando {rows_to_delete_count} filas...")
                sheet.delete_rows(2, amount=rows_to_delete_count)
                first_data_row_actual_index = 2
            elif first_data_row_index == -1:
                _logger.warning("No se pudo encontrar la línea de inicio del cuerpo...")
            # --- FIN SECCIÓN 4 MEJORADA ---

            # --- INICIO MODIFICACIÓN (REQUISITOS EXCEL CORREGIDOS - NUEVO ENFOQUE OCULTAR CEROS) ---
            # --- 4.5. DETECTAR NIVELES Y APLICAR LÓGICA DE FILAS (OCULTAR 0/VACÍO, INSERTAR TÍTULO, NEGRITA N1) ---
            _logger.info(f"--- Iniciando lógica de formato (Negritas N1, Ocultar Ceros/Vacíos) e Inserción de Filas (Títulos) ---")
            base_indent = 0
            level_1_indent = 2
            try:
                first_cell_alignment = sheet.cell(row=first_data_row_actual_index, column=1).alignment
                if first_cell_alignment and first_cell_alignment.indent is not None:
                    base_indent = first_cell_alignment.indent
                else:
                     _logger.warning(f"No se pudo obtener indentación para la celda A{first_data_row_actual_index}, usando base_indent=0.")
                level_1_indent = base_indent + 2
                _logger.info(f"Indent base detectado: {base_indent}. Usando Nivel 1={level_1_indent}")
            except Exception as e:
                _logger.warning(f"Excepción al detectar indent base, usando valores por defecto. Error: {e}")

            ZERO_THRESHOLD = 0.005 # Umbral para considerar cero
            rows_to_process = []

            _logger.debug("--- Recopilando información de filas (de abajo hacia arriba) ---")
            for row_idx in range(sheet.max_row, first_data_row_actual_index - 1, -1):
                try:
                    row = sheet[row_idx]
                    # Validación más robusta de fila vacía
                    if not any(cell.value is not None and str(cell.value).strip() != '' for cell in row):
                         _logger.debug(f"Saltando fila {row_idx} por estar completamente vacía.")
                         continue # Ignorar filas completamente vacías
                    cell_A = row[0]
                    # Asegurarse de que cell_A tenga valor antes de proceder
                    if cell_A.value is None or str(cell_A.value).strip() == '':
                         _logger.debug(f"Saltando fila {row_idx} por celda A vacía.")
                         continue
                    row_indent = cell_A.alignment.indent if cell_A.alignment and cell_A.alignment.indent is not None else 0
                    rows_to_process.append({'idx': row_idx, 'row': row, 'indent': row_indent})
                except Exception as e:
                    _logger.error(f"Error al recopilar información de fila {row_idx}: {e}", exc_info=True)

            _logger.debug(f"--- Procesando {len(rows_to_process)} filas recopiladas ---")
            processed_indices = set()
            indices_to_hide = set() # Guardar índices de filas a ocultar

            # Primera pasada: Identificar filas a ocultar y aplicar formato N1/Insertar filas Título
            for row_data in reversed(rows_to_process):
                row_idx = row_data['idx']
                if row_idx in processed_indices: continue

                row = row_data['row']
                cell_A = row[0]
                row_indent = row_data['indent']
                current_processing_idx = row_idx # Índice original

                try:
                    # --- REQUISITO 1: FILA EN BLANCO ANTES DE TÍTULOS ESPECIALES (NIVEL 0) ---
                    if row_indent == base_indent and isinstance(cell_A.value, str):
                        cell_A_text_upper = cell_A.value.strip().upper()
                        if cell_A_text_upper in SPECIAL_HEADERS_TUPLE:
                            # Marcar para insertar DESPUÉS, pero registrar el índice ORIGINAL
                            # No insertamos aquí para no complicar los índices
                            _logger.debug(f"Marcar para insertar fila antes de '{cell_A.value}' (fila original {row_idx})")
                            # La inserción se hará en una pasada posterior

                    is_level_1 = (row_indent == level_1_indent)

                    # --- REQUISITO 3: IDENTIFICAR FILAS CON VALOR ~0.00 o VACÍAS PARA OCULTAR ---
                    # (No ocultar Nivel 0 ni Nivel 1)
                    if row_indent > level_1_indent:
                        data_cells = row[1:]
                        all_effectively_zero_or_empty = True
                        has_numeric_or_text_in_data = False # Verificar si hay ALGO en las columnas de datos

                        for cell in data_cells:
                            if isinstance(cell.value, (int, float)):
                                has_numeric_or_text_in_data = True
                                if abs(cell.value) >= ZERO_THRESHOLD:
                                    all_effectively_zero_or_empty = False
                                    break
                            elif isinstance(cell.value, str) and cell.value.strip():
                                has_numeric_or_text_in_data = True
                                all_effectively_zero_or_empty = False
                                break
                            # None o '' se consideran vacíos/cero

                        # Marcar para ocultar SI tiene algo (para no ocultar separadores vacíos)
                        # Y si todo era cero/vacío
                        if has_numeric_or_text_in_data and all_effectively_zero_or_empty:
                            indices_to_hide.add(current_processing_idx)
                            _logger.debug(f"Marcando fila {current_processing_idx} para ocultar (valor cero o vacío): '{cell_A.value}'")
                            processed_indices.add(row_idx)
                            continue # Marcada para ocultar, no necesita formato N1

                    # --- REQUISITO 2: NEGRITA EN LÍNEAS DE NIVEL 1 ---
                    # (Aplicar solo si la fila no fue marcada para ocultar)
                    if is_level_1:
                        # Aplicar directamente ya que no afecta índices
                        current_row_obj = sheet[current_processing_idx] # Obtener la fila actual por si acaso
                        for cell in current_row_obj:
                            if cell.value is not None and str(cell.value).strip() != '':
                                cell.font = bold_data_font

                    processed_indices.add(row_idx) # Marcar como procesado (formato o chequeo hecho)

                except Exception as e:
                    _logger.error(f"Error CRÍTICO al procesar fila índice original {row_idx}: {e}", exc_info=True)
                    processed_indices.add(row_idx)

            # Segunda pasada: Aplicar ocultación e inserción de abajo hacia arriba
            _logger.debug(f"--- Aplicando ocultaciones e inserciones (de abajo hacia arriba) ---")
            sorted_indices_to_hide = sorted(list(indices_to_hide), reverse=True)
            for idx in sorted_indices_to_hide:
                 try:
                     if idx not in sheet.row_dimensions:
                         sheet.row_dimensions[idx] = openpyxl.worksheet.dimensions.RowDimension(sheet, index=idx)
                     sheet.row_dimensions[idx].hidden = True
                     _logger.debug(f"Ocultando fila índice {idx}")
                 except Exception as e:
                     _logger.error(f"Error al ocultar fila {idx}: {e}", exc_info=True)

            # Re-iterar para insertar filas antes de Títulos (de abajo hacia arriba)
            for row_data in reversed(rows_to_process):
                 row_idx = row_data['idx']
                 # Solo procesar las que no estaban marcadas para ocultar
                 if row_idx not in indices_to_hide:
                     cell_A = row_data['row'][0]
                     row_indent = row_data['indent']
                     if row_indent == base_indent and isinstance(cell_A.value, str):
                         cell_A_text_upper = cell_A.value.strip().upper()
                         if cell_A_text_upper in SPECIAL_HEADERS_TUPLE:
                             try:
                                 sheet.insert_rows(row_idx)
                                 _logger.debug(f"Insertando fila antes de Título en índice original {row_idx}")
                             except Exception as e:
                                 _logger.error(f"Error al insertar fila antes de índice {row_idx}: {e}", exc_info=True)


            _logger.info(f"--- Fin de formato y lógica de filas Excel ---")

            # --- 5. APLICAR FORMATO DE NÚMERO (Separado) ---
            # (Se mantiene igual, solo itera sobre filas visibles)
            _logger.info(f"--- Aplicando formato de número personalizado a la hoja ---")
            number_format_str = '#,##0.00 ;[Red](#,##0.00)'
            for row_idx in range(1, sheet.max_row + 1):
                if row_idx == 1 and header_text: continue
                # Verificar si la fila existe y está oculta
                is_hidden = row_idx in sheet.row_dimensions and sheet.row_dimensions[row_idx].hidden
                if is_hidden or not sheet[row_idx] or not any(c.value is not None for c in sheet[row_idx]): # Saltar ocultas o vacías
                     continue

                for cell in sheet[row_idx]:
                    if isinstance(cell.value, (int, float)):
                        if cell.value is not None:
                           cell.number_format = number_format_str
            _logger.info(f"--- Formato de número aplicado ---")
            # --- FIN MODIFICACIÓN ---

            # =========================================================================
            # --- 6. AGREGAR CERTIFICACIÓN Y FIRMAS (NUEVA LÓGICA HORIZONTAL) ---
            # =========================================================================
            # (Esta sección no cambia)
            # (...) Código de firmas omitido por brevedad
            current_row = sheet.max_row + 3
            max_col = sheet.max_column if sheet.max_column >= 1 else 1
            max_col_letter = get_column_letter(max_col)
            # --- 6.1. Texto de Certificación ---
            if company.certification_text:
                merge_range_cert = f'A{current_row}:{max_col_letter}{current_row + 1}' if max_col >= 1 else f'A{current_row}'
                try: sheet.merge_cells(merge_range_cert)
                except Exception as merge_err: _logger.error(f"Error al combinar celdas para certificación '{merge_range_cert}': {merge_err}.")
                cell = sheet[f'A{current_row}']
                cell.value = company.certification_text; cell.font = normal_font; cell.alignment = left_align
                if current_row not in sheet.row_dimensions or sheet.row_dimensions[current_row].height is None:
                     # Estimar altura si no existe dimensión
                     estimated_height = max(15, len(company.certification_text.split('\n')) * 15)
                     sheet.row_dimensions[current_row].height = estimated_height
                else:
                    sheet.row_dimensions[current_row].height = max(15, len(company.certification_text.split('\n')) * 15)

                current_row += 3
            # --- 6.2. Helper Interno ---
            def _write_signature_block(sheet, start_row, col_start, col_end, name, title):
                 # (...) Implementación sin cambios
                _logger.info(f"Escribiendo firma '{name}' de C{col_start} a C{col_end} en fila {start_row}")
                try:
                    line_row, name_row, title_row = start_row, start_row + 1, start_row + 2
                    start_letter, end_letter = get_column_letter(col_start), get_column_letter(col_end)
                    can_merge = col_end > col_start
                    # Línea
                    merge_range_line = f'{start_letter}{line_row}:{end_letter}{line_row}' if can_merge else f'{start_letter}{line_row}'
                    if can_merge: sheet.merge_cells(merge_range_line)
                    cell_line = sheet.cell(row=line_row, column=col_start, value="___________________________")
                    cell_line.alignment = sign_center_align; cell_line.font = normal_font
                    if line_row not in sheet.row_dimensions or sheet.row_dimensions[line_row].height is None: sheet.row_dimensions[line_row].height = 18
                    # Nombre
                    merge_range_name = f'{start_letter}{name_row}:{end_letter}{name_row}' if can_merge else f'{start_letter}{name_row}'
                    if can_merge: sheet.merge_cells(merge_range_name)
                    cell_name = sheet.cell(row=name_row, column=col_start, value=name)
                    cell_name.alignment = sign_center_align; cell_name.font = bold_font
                    if name_row not in sheet.row_dimensions or sheet.row_dimensions[name_row].height is None: sheet.row_dimensions[name_row].height = 18
                     # Título
                    merge_range_title = f'{start_letter}{title_row}:{end_letter}{title_row}' if can_merge else f'{start_letter}{title_row}'
                    if can_merge: sheet.merge_cells(merge_range_title)
                    cell_title = sheet.cell(row=title_row, column=col_start, value=title)
                    cell_title.alignment = sign_center_align; cell_title.font = normal_font
                    if title_row not in sheet.row_dimensions or sheet.row_dimensions[title_row].height is None: sheet.row_dimensions[title_row].height = 15
                    return title_row
                except Exception as merge_err:
                    _logger.error(f"Error al procesar/combinar celdas para firma '{name}': {merge_err}")
                    sheet.cell(row=start_row, column=col_start).value = "___________________________"
                    sheet.cell(row=start_row + 1, column=col_start).value = name
                    sheet.cell(row=start_row + 2, column=col_start).value = title
                    return start_row + 2
            # --- 6.3. Lógica de Firmas ---
            signatures = []
            if company.name_1 and company.position_1: signatures.append((company.name_1, company.position_1))
            if company.name_2 and company.position_2: signatures.append((company.name_2, company.position_2))
            if company.name_3 and company.position_3: signatures.append((company.name_3, company.position_3))
            num_signatures = len(signatures)
            _logger.info(f"Iniciando escritura de {num_signatures} firmas en layout horizontal.")
            if num_signatures > 0:
                if num_signatures == 1: name, title = signatures[0]; _write_signature_block(sheet, current_row, 1, max_col, name, title)
                elif num_signatures == 2:
                    mid_point = max(1, max_col // 2)
                    name1, title1 = signatures[0]; _write_signature_block(sheet, current_row, 1, mid_point, name1, title1)
                    name2, title2 = signatures[1]; _write_signature_block(sheet, current_row, mid_point + 1, max_col, name2, title2)
                elif num_signatures >= 3:
                    mid_point = max(1, max_col // 2)
                    name1, title1 = signatures[0]; _write_signature_block(sheet, current_row, 1, mid_point, name1, title1)
                    name2, title2 = signatures[1]; _write_signature_block(sheet, current_row, mid_point + 1, max_col, name2, title2)
                    current_row += 5
                    name3, title3 = signatures[2]; _write_signature_block(sheet, current_row, 1, max_col, name3, title3)


            # --- 7. GUARDAR Y DEVOLVER EL ARCHIVO MODIFICADO ---
            new_output = io.BytesIO()
            workbook.save(new_output)
            new_output.seek(0)
            modified_bytes = new_output.read()

            original_export_data['file_content'] = modified_bytes
            return original_export_data

        except Exception as e:
            _logger.error(f"Error CRÍTICO al modificar el archivo XLSX (reporte: {self.name}): {e}", exc_info=True)
            # Devolver el original en caso de error grave
            original_export_data['file_content'] = original_output_bytes
            return original_export_data